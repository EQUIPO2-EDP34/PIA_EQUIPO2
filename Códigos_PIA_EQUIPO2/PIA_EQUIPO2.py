import random as rd
import sys
from datetime import datetime
import sqlite3
from sqlite3 import Error
import openpyxl

def crear_tablas():
    try:
        with sqlite3.connect("BD_NISSAN.db") as conn:
            mi_cursor = conn.cursor()

            # Creamos la tabla Sucursales
            mi_cursor.execute("""
            CREATE TABLE IF NOT EXISTS Sucursales (
                ID_Sucursal INTEGER NOT NULL PRIMARY KEY,
                Nombre TEXT,
                Direccion TEXT,
                Teléfono INTEGER
            )
            """)

            # Creamos la tabla Autos
            mi_cursor.execute("""
            CREATE TABLE IF NOT EXISTS Autos (
                ID_AUTO INTEGER NOT NULL PRIMARY KEY,
                MODELO TEXT,
                TIPO_AUTO TEXT,
                PRECIO REAL,
                ANIO INTEGER,
                TIPO_COMBUSTIBLE TEXT,
                LITROS_MOTOR REAL,
                COLOR TEXT,
                STOCK INTEGER
            )
            """)
            #Creamos la tabla Ventas
            mi_cursor.execute("""
            CREATE TABLE IF NOT EXISTS Ventas (
            ID_Venta INTEGER NOT NULL PRIMARY KEY,
            ID_auto INTEGER,
            ID_sucursal INTEGER,
            Nombre_auto TEXT,
            Nombre_sucursal TEXT,
            Nombre_cliente TEXT, 
            Cantidad INTEGER,
            Fecha DATE,
            FOREIGN KEY (ID_auto) REFERENCES Autos(ID_AUTO),
            FOREIGN KEY (ID_sucursal) REFERENCES Sucursales(ID_Sucursal)
                )
                """)    

            print("Tablas creadas exitosamente")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

#opcion 1
def registrar_venta():
    try:
        while True:
            print("***¡Tome en cuenta que debe existir registro de auto y sucursal para poder registrar una venta!***")
            auto_id = int(input("Ingrese ID del auto: "))
            sucursal_id = int(input("Ingrese ID de la sucursal: "))
            nombre_cliente = input("Ingrese el nombre del cliente: ")
            cantidad = int(input("Ingrese cantidad vendida: "))

            # Solicitar fecha en formato DD-MM-YYYY
            fecha_str = input("Ingrese la fecha de la venta (DD-MM-YYYY): ")

            try:
                # Convertir la cadena de fecha a objeto datetime
                fecha = datetime.strptime(fecha_str, "%d-%m-%Y").date()
            except ValueError:
                print("Formato de fecha incorrecto. Por favor, ingrese la fecha en formato DD-MM-YYYY.")
                continue

            # Verificar si el ID del auto existe en la tabla Autos
            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT * FROM Autos WHERE ID_AUTO = ?", (auto_id,))
                auto_existente = mi_cursor.fetchone()

                if not auto_existente:
                    print("El ID del auto ingresado no existe. Por favor, verifique e intente nuevamente.")
                    respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
                    if respuesta.lower() == 'salir':
                        break
                    continue

            # Verificar si el ID de la sucursal existe en la tabla Sucursales
            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT * FROM Sucursales WHERE ID_Sucursal = ?", (sucursal_id,))
                sucursal_existente = mi_cursor.fetchone()

                if not sucursal_existente:
                    print("El ID de la sucursal ingresado no existe. Por favor, verifique e intente nuevamente.")
                    respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
                    if respuesta.lower() == 'salir':
                        break
                    continue

            # Si ambos IDs existen, realizar la inserción en la tabla Ventas
            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                valores = (auto_id, sucursal_id, nombre_cliente, cantidad, fecha)
                mi_cursor.execute("INSERT INTO Ventas (ID_auto, ID_sucursal, Nombre_cliente, Cantidad, Fecha) VALUES (?, ?, ?, ?, ?)", valores)

                # Obtener el ID de la venta recién registrada
                mi_cursor.execute("SELECT last_insert_rowid()")
                id_venta = mi_cursor.fetchone()[0]

                print(f"Venta registrada exitosamente. ID de la venta: {id_venta}")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break

    except Error as e:
        print(e)

#opcion 2
def modificar_venta():
    try:
        while True:
            venta_id = int(input("Ingrese ID de la venta a modificar: "))
            nueva_cantidad = int(input("Ingrese nueva cantidad vendida: "))
            nuevo_nombre_cliente = input("Ingrese el nuevo nombre del cliente: ")
            nueva_fecha_str = input("Ingrese la nueva fecha de la venta (DD-MM-YYYY): ")

            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()

                # Obtenemos la venta actual para mostrar los detalles antes de la modificación
                mi_cursor.execute("""
                    SELECT Ventas.ID_Venta, Autos.ID_AUTO as ID_auto, Autos.MODELO as Nombre_auto,
                           Sucursales.ID_Sucursal as ID_sucursal, Sucursales.Nombre as Nombre_sucursal,
                           Ventas.Nombre_cliente, Ventas.Cantidad, Ventas.Fecha
                    FROM Ventas
                    JOIN Autos ON Ventas.ID_auto = Autos.ID_AUTO
                    JOIN Sucursales ON Ventas.ID_sucursal = Sucursales.ID_Sucursal
                    WHERE Ventas.ID_Venta = ?
                """, (venta_id,))
                venta_actual = mi_cursor.fetchone()

                if not venta_actual:
                    print("No se encontró una venta con el ID proporcionado. Verifique el ID e intente nuevamente.")
                    respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
                    if respuesta.lower() == 'salir':
                        break
                    continue

                print("\nDetalles de la venta actual:")
                print(f"ID_Venta: {venta_actual[0]}")
                print(f"ID_auto: {venta_actual[1]}, Nombre_auto: {venta_actual[2]}")
                print(f"ID_sucursal: {venta_actual[3]}, Nombre_sucursal: {venta_actual[4]}")
                print(f"Nombre_cliente actual: {venta_actual[5]}")
                print(f"Cantidad actual: {venta_actual[6]}")
                print(f"Fecha actual: {venta_actual[7]}")

                # Modificamos la venta con los nuevos datos
                valores = (nueva_cantidad, nuevo_nombre_cliente, nueva_fecha_str, venta_id)
                mi_cursor.execute("""
                    UPDATE Ventas
                    SET Cantidad = ?,
                        Nombre_cliente = COALESCE(?, Nombre_cliente),  -- Mantener el valor actual si se deja en blanco
                        Fecha = COALESCE(?, Fecha)  -- Mantener el valor actual si se deja en blanco
                    WHERE ID_Venta = ?
                """, valores)

                print("Venta modificada exitosamente")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break

    except Error as e:
        print(e)

#opcion 3
def registrar_sucursal():
    try:
        while True:
            nombre = input("Ingrese el nombre de la sucursal: ")
            direccion = input("Ingrese la dirección de la sucursal: ")
            telefono = input("Ingrese el teléfono de la sucursal: ")

            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                valores = (nombre, direccion, telefono)
                mi_cursor.execute("INSERT INTO Sucursales (Nombre, Direccion, Teléfono) VALUES (?, ?, ?)", valores)

            # Obtenemos el ID de la sucursal recién registrada
                mi_cursor.execute("SELECT last_insert_rowid()")
                id_sucursal = mi_cursor.fetchone()[0]

                print(f"Sucursal registrada exitosamente. ID de la sucursal: {id_sucursal}")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break
    except Error as e:
        print(e)

#opcion 4
def registrar_auto():
    try:
        while True:
            modelo = input("Ingrese el modelo del auto: ")
            tipo_auto = input("Ingrese el tipo de auto: ")
            precio = float(input("Ingrese el precio del auto: "))
            anio = int(input("Ingrese el año del auto: "))
            tipo_combustible = input("Ingrese el tipo de combustible del auto: ")
            litros_motor = float(input("Ingrese los litros del motor del auto: "))
            color = input("Ingrese el color del auto: ")
            stock = int(input("Ingrese el stock inicial del auto: "))

            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                valores = (modelo, tipo_auto, precio, anio, tipo_combustible, litros_motor, color, stock)
                mi_cursor.execute("INSERT INTO Autos (MODELO, TIPO_AUTO, PRECIO, ANIO, TIPO_COMBUSTIBLE, LITROS_MOTOR, COLOR, STOCK) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", valores)

            # Obtenemos el ID del auto recién registrado
                mi_cursor.execute("SELECT last_insert_rowid()")
                id_auto = mi_cursor.fetchone()[0]

                print(f"Auto registrado exitosamente. ID del auto: {id_auto}")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break
    except Error as e:
        print(e)

#opcion 5
def eliminar_venta():
    try:
        while True:
            venta_id = int(input("Ingrese ID de la venta a eliminar: "))

            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()

                # Verificamos si la venta con el ID proporcionado existe
                mi_cursor.execute("SELECT * FROM Ventas WHERE ID_Venta = ?", (venta_id,))
                venta_existente = mi_cursor.fetchone()

                if not venta_existente:
                    print("No se encontró una venta con el ID proporcionado. Verifique el ID e intente nuevamente.")
                    respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
                    if respuesta.lower() == 'salir':
                        break
                    continue

                # Eliminamos la venta con el ID proporcionado
                mi_cursor.execute("DELETE FROM Ventas WHERE ID_Venta = ?", (venta_id,))
                print("Venta eliminada exitosamente")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break

    except Error as e:
        print(e)

#opcion 6
def mostrar_autos():
    try:
        with sqlite3.connect("BD_NISSAN.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Autos")
            autos = mi_cursor.fetchall()

            if not autos:
                print("No hay autos registrados en la base de datos.")
            else:
                print("Autos registrados:")
                for auto in autos:
                    print(f"ID: {auto[0]}, Modelo: {auto[1]}, Tipo: {auto[2]}, Precio: {auto[3]}, Año: {auto[4]}, Combustible: {auto[5]}, Litros: {auto[6]}, Color: {auto[7]}, Stock: {auto[8]}")
            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                return
    except Error as e:
        print(e)

#opcion 7
def mostrar_sucursales():
    try:
        with sqlite3.connect("BD_NISSAN.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Sucursales")
            sucursales = mi_cursor.fetchall()

            if not sucursales:
                print("No hay sucursales registradas en la base de datos.")
            else:
                print("Sucursales registradas:")
                for sucursal in sucursales:
                    print(f"ID: {sucursal[0]}, Nombre: {sucursal[1]}, Dirección: {sucursal[2]}, Teléfono: {sucursal[3]}")
            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                return
    except Error as e:
        print(e)

#opcion 8
def mostrar_ventas():
    try:
        with sqlite3.connect("BD_NISSAN.db") as conn:
            mi_cursor = conn.cursor()

            opcion_mostrar_todas = input("¿Quiere ver todas las ventas? \nResponda 'si' o 'no' para ver solo una venta en especifico): ").lower()

            if opcion_mostrar_todas.upper() == 'SI':
                mi_cursor.execute("""
                    SELECT Ventas.ID_Venta, Autos.ID_AUTO as ID_auto, Autos.MODELO as Nombre_auto,
                           Sucursales.ID_Sucursal as ID_sucursal, Sucursales.Nombre as Nombre_sucursal,
                           Ventas.Nombre_cliente, Ventas.Cantidad, Ventas.Fecha
                    FROM Ventas
                    JOIN Autos ON Ventas.ID_auto = Autos.ID_AUTO
                    JOIN Sucursales ON Ventas.ID_sucursal = Sucursales.ID_Sucursal
                """)
            elif opcion_mostrar_todas.upper() == 'NO':
                while True:
                    venta_id_str = input("Ingrese el ID de la venta que desea ver: ")

                    if not venta_id_str:
                        print("Saliendo...")
                        return

                    try:
                        venta_id = int(venta_id_str)
                        break
                    except ValueError:
                        print("Por favor, ingrese un número válido.")

                mi_cursor.execute("""
                    SELECT Ventas.ID_Venta, Autos.ID_AUTO as ID_auto, Autos.MODELO as Nombre_auto,
                           Sucursales.ID_Sucursal as ID_sucursal, Sucursales.Nombre as Nombre_sucursal,
                           Ventas.Nombre_cliente, Ventas.Cantidad, Ventas.Fecha
                    FROM Ventas
                    JOIN Autos ON Ventas.ID_auto = Autos.ID_AUTO
                    JOIN Sucursales ON Ventas.ID_sucursal = Sucursales.ID_Sucursal
                    WHERE Ventas.ID_Venta = ?
                """, (venta_id,))
            else:
                print("Respuesta no válida. Volviendo al menú principal.")
                return

            ventas = mi_cursor.fetchall()

            if not ventas:
                print("No hay ventas registradas en la base de datos.")
            else:
                print("*****Ventas registradas: *****")
                for venta in ventas:
                    print(f"ID_Venta: {venta[0]}, ID_auto: {venta[1]}, Nombre_auto: {venta[2]}, "
                          f"ID_sucursal: {venta[3]}, Nombre_sucursal: {venta[4]}, "
                          f"Nombre_cliente: {venta[5]}, Cantidad: {venta[6]}, Fecha: {venta[7]}")

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                return
    except Error as e:
        print(e)

#opcion 9
def exportar_base_de_datos_a_excel():
    try:
        while True:
            with sqlite3.connect("BD_NISSAN.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("""
                    SELECT Ventas.ID_Venta, Autos.ID_AUTO as ID_auto, Autos.MODELO as Nombre_auto,
                           Sucursales.ID_Sucursal as ID_sucursal, Sucursales.Nombre as Nombre_sucursal,
                           Ventas.Nombre_cliente, Ventas.Cantidad, Ventas.Fecha
                    FROM Ventas
                    JOIN Autos ON Ventas.ID_auto = Autos.ID_AUTO
                    JOIN Sucursales ON Ventas.ID_sucursal = Sucursales.ID_Sucursal
                """)
                registros = mi_cursor.fetchall()

                wb = openpyxl.Workbook()
                hoja = wb.active

                columnas = ["ID_Venta", "ID_auto", "Nombre_auto", "ID_sucursal", "Nombre_sucursal", "Nombre_cliente", "Cantidad", "Fecha"]

                for i, column_name in enumerate(columnas, start=1):
                    hoja.cell(row=1, column=i, value=column_name)

                for row, record in enumerate(registros, start=2):
                    for col, value in enumerate(record, start=1):
                        hoja.cell(row=row, column=col, value=value)

                wb.save('exportacion_ventas.xlsx')
                print('Datos exportados a exportacion_ventas.xlsx')

            respuesta = input("¿Desea regresar al menú principal? (responda 'salir' para salir): ")
            if respuesta.lower() == 'salir':
                break
    except Error as e:
        print(e)

def mostrar_menu_principal():
    print("*" * 20)
    print("1. Registrar Venta\n"
          "2. Modificar Venta\n"
          "3. Registrar Sucursal\n"
          "4. Registrar Auto\n"
          "5. Eliminar Venta\n"
          "6. Mostrar Autos\n"
          "7. Mostrar Sucursales\n" 
          "8. Mostrar Ventas\n"
          "9. Exportar Base de Datos a Excel\n"
          "10. Salir del programa\n")

def mi_menu():
    print("Conexion Establecida")
    crear_tablas()
    
    while True:
        mostrar_menu_principal()
        
        try:
            opcion = int(input("Seleccione el número de la acción que quiere realizar: "))
        except ValueError:
            print("Por favor, ingrese un número válido.")
            continue

        if opcion == 1:
            registrar_venta()
        elif opcion == 2:
            modificar_venta()
        elif opcion == 3:
            registrar_sucursal()
        elif opcion == 4:
            registrar_auto()
        elif opcion == 5:
            eliminar_venta()
        elif opcion == 6:
            mostrar_autos()
        elif opcion == 7:
            mostrar_sucursales() 
        elif opcion == 8:
            mostrar_ventas()
        elif opcion == 9:
            exportar_base_de_datos_a_excel()
        elif opcion == 10:
            print("Hasta pronto. Ten un buen día :D")
            break
        else:
            print("Esa opción no está disponible. Por favor, revisa el menú.") 

if __name__ == "__main__":
    mi_menu()